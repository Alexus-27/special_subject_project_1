import openpyxl
import sys
import os
import re
from PyQt5.QtGui import QTextCharFormat, QTextCursor, QColor, QFont
from PyQt5.QtCore import QTimer, Qt, QPoint
from PyQt5.QtWidgets import (
    QApplication, QFileDialog, QFrame, QLineEdit, QMessageBox, 
    QProgressBar, QTableWidget, QTableWidgetItem, QWidget, QVBoxLayout, 
    QHBoxLayout, QPushButton, QLabel, QSizePolicy, QTextEdit, QScrollArea,
    QHeaderView
)

class CustomHeader(QFrame):
    def __init__(self, parent, active=None):
        super().__init__(parent)
        self.parent = parent
        self.setFixedHeight(40)
        self.setStyleSheet("""
            QFrame {
                background-color: lightgray;
                border-bottom: 2px solid black;
            }
        """)

        layout = QHBoxLayout()
        layout.setContentsMargins(5, 2, 5, 2)
        layout.setSpacing(5)

        # Left-side app buttons
        self.search_btn = QPushButton("Search File")
        self.help_btn = QPushButton("Help")
        self.about_btn = QPushButton("About")

        for btn in [self.search_btn, self.help_btn, self.about_btn]:
            btn.setFixedSize(100, 30)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: white;
                    border: 1px solid black;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e0e0e0;
                }
            """)

        layout.addWidget(self.search_btn)
        layout.addWidget(self.help_btn)
        layout.addWidget(self.about_btn)
        layout.addStretch()

        # Window control buttons
        self.minimize_btn = QPushButton("–")
        self.maximize_btn = QPushButton("🗖")

        self.close_btn = QPushButton("X")

        # self.maximize_btn.setText("□")
        # self.maximize_btn.setToolTip("Maximize")

        # self.minimize_btn.setText("–")
        # self.maximize_btn.setToolTip("Minimize")

        

        for btn in [self.minimize_btn, self.maximize_btn, self.close_btn]:
            btn.setFixedSize(30, 30)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: white;
                    border: 1px solid black;
                }
                QPushButton:hover {
                    background-color: #f0f0f0;
                }
            """)

        layout.addWidget(self.minimize_btn)
        layout.addWidget(self.maximize_btn)
        layout.addWidget(self.close_btn)

        self.setLayout(layout)

        # Set initial state of maximize button depending on BaseWindow.window_is_maximized
        if BaseWindow.window_is_maximized:
            self.maximize_btn.setText("🗗")  # Or 🗗
            self.maximize_btn.setToolTip("Restore")
        else:
            self.maximize_btn.setText("🗖")
            self.maximize_btn.setToolTip("Maximize")


        # Connections
        self.close_btn.clicked.connect(QApplication.quit)
        self.minimize_btn.clicked.connect(parent.showMinimized)
        self.maximize_btn.clicked.connect(self.toggle_max_restore)

        # Active button has a color 
        self.active_btn = None
        self.search_btn.clicked.connect(lambda: self.handle_active(self.search_btn, parent.open_search_window))
        self.help_btn.clicked.connect(lambda: self.handle_active(self.help_btn, parent.open_help_window))
        self.about_btn.clicked.connect(lambda: self.handle_active(self.about_btn, parent.open_about_window))

        # Color for active button 
        if active == "Search":
            self.set_active_button(self.search_btn)
        elif active == "Help":
            self.set_active_button(self.help_btn)
        elif active == "About":
            self.set_active_button(self.about_btn)


    def toggle_max_restore(self):
        if self.parent.isMaximized():
            self.parent.showNormal()
            BaseWindow.window_is_maximized = False
            self.maximize_btn.setText("🗖")
            self.maximize_btn.setToolTip("Maximize")
        else:
            self.parent.showMaximized()
            BaseWindow.window_is_maximized = True
            self.maximize_btn.setText("🗗")
            self.maximize_btn.setToolTip("Minimize")


    def handle_active(self, button, action):
        self.set_active_button(button)
        action()

    def set_active_button(self, button):
        if self.active_btn:
            self.active_btn.setStyleSheet(self.button_style())  # update style of the button

        button.setStyleSheet(self.button_style(active=True))
        self.active_btn = button

    def button_style(self, active=False):
        base = """
            QPushButton {{
                background-color: {bg};
                border: 2px solid black;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {hover};
            }}
        """
        if active:
            return base.format(bg="#d0d0ff", hover="#c0c0ff")  # active is light-blue color
        else:
            return base.format(bg="white", hover="#e0e0e0")



class BaseWindow(QWidget):
    shared_df = None
    window_is_maximized = False
    latest_version = -1
    def __init__(self):
        super().__init__()
        self.old_pos = None
        self.selected_files = []

        if BaseWindow.shared_df is None:
            try:
                base_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
                mapping_dir = os.path.join(base_dir, "data")
                latest_file = None

                for filename in os.listdir(mapping_dir):
                    if filename.startswith("mapping_table") and filename.endswith(".xlsx"):
                        match = re.search(r"mapping_table(\d+)\.xlsx", filename)
                        if match:
                            version = int(match.group(1))
                            if version > BaseWindow.latest_version:
                                BaseWindow.latest_version = version
                                latest_file = filename

                if latest_file:
                    file_path = os.path.join(mapping_dir, latest_file)
                    workbook = openpyxl.load_workbook(file_path)
                    sheet = workbook["Sample"]

                    # take headlines
                    headers = [cell.value.strip() if isinstance(cell.value, str) else str(cell.value)
                               for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

                    # Read lines with logic ffil
                    data = []
                    previous_cause = None
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        row_dict = dict(zip(headers, row))
                        
                        # Types
                        code = str(row_dict.get("Err Code")).strip() if row_dict.get("Err Code") is not None else ""
                        cause = row_dict.get("Cause")
                        action = str(row_dict.get("Action")).strip() if row_dict.get("Action") is not None else ""

                        # Enter Cause from previous, if empty
                        if cause is None:
                            cause = previous_cause
                        else:
                            previous_cause = cause

                        data.append({
                            "Err Code": code,
                            "Cause": cause,
                            "Action": action
                        })


                    BaseWindow.shared_df = data
                    print(f"[INFO] Loaded mapping table: {latest_file}")
                else:
                    raise FileNotFoundError("No valid mapping_tableXXXX.xlsx file found in /data.")

            except Exception as e:
                print(f"[ERROR] Failed to load mapping table: {e}")

        self.df = BaseWindow.shared_df

    # Moving for Window
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.old_pos = event.globalPos()

    def mouseMoveEvent(self, event):
        if self.old_pos:
            delta = QPoint(event.globalPos() - self.old_pos)
            self.move(self.x() + delta.x(), self.y() + delta.y())
            self.old_pos = event.globalPos()

    def mouseReleaseEvent(self, event):
        self.old_pos = None

    # Header functions for buttons
    def open_search_window(self):
    # Open dilog after choosing file or files
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select one or more log files",
            "",
            "Log Files (*.log);;All Files (*)"
        )

        if files:
            # Filter only log files
            valid_files = [f for f in files if f.lower().endswith(".log")]

            if not valid_files:
                QMessageBox.warning(self, "Invalid File(s)", "Please select only .log files.")
                return

            # Check for empty
            non_empty_files = []
            for f in valid_files:
                try:
                    if os.path.getsize(f) > 0:
                        non_empty_files.append(f)
                except Exception as e:
                    print(f"[WARNING] Could not check file {f}: {e}")

            # If one have chosen and it's empty
            if len(valid_files) == 1 and not non_empty_files:
                QMessageBox.warning(self, "Empty File", "No content found. Please, select non-empty file!")
                return

            if not non_empty_files:
                QMessageBox.warning(self, "No Valid Files", "No non-empty log files found.")
                return

            self.selected_files = non_empty_files
            self.user_choice_window = UserChoiceWindow(non_empty_files)
            self.user_choice_window.show()
            self.close()
        else:
            self.selected_files = []


    def open_help_window(self):
        self.help_window = HelpWindow()
        self.help_window.show()
        self.close()

    def open_about_window(self):
        self.about_window = AboutWindow()
        self.about_window.show()
        self.close()

    def show_user_choice(self, parent_dialog):
        parent_dialog.accept()
        QMessageBox.information(self, "User's Choice", "You selected a file!")

    @staticmethod
    def button_style(base_color="white", hover_color="#e0e0e0", border="1px solid black", font_size="14px", bold=False):
        return f"""
            QPushButton {{
                background-color: {base_color};
                border: {border};
                font-size: {font_size};
                {"font-weight: bold;" if bold else ""}
            }}
            QPushButton:hover {{
                background-color: {hover_color};
            }}
        """

class MainWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setMinimumSize(800, 600)
        self.selected_files = []  # selected files (logs)

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        self.header = CustomHeader(self)
        layout.addWidget(self.header)

        # Centered container of instuctions
        center_container = QWidget()
        center_layout = QVBoxLayout(center_container)
        center_layout.setAlignment(Qt.AlignCenter)

        panel = QFrame()
        panel.setStyleSheet("background-color: #f0f0f0; border: 1px solid #999999; border-radius: 8px;")
        panel.setFixedSize(620, 260)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(25, 20, 25, 20)
        panel_layout.setSpacing(12)

        # Header
        title = QLabel("Instructions")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 18px; font-weight: bold; color: #333333;")
        panel_layout.addWidget(title)

        # Instructions
        instructions = [
            "1. 'Search File' button opens the user directory to choose a file or files for analyzing.",
            "2. 'Help' button opens a new window where user can search error codes to find reasons and corrective actions.",
            "3. 'About' button shows the main information about system and mapping table's version."
        ]

        for text in instructions:
            label = QLabel(text)
            label.setWordWrap(True)
            label.setAlignment(Qt.AlignLeft)
            label.setStyleSheet("font-size: 14px; color: #222222;")
            panel_layout.addWidget(label)

        center_layout.addWidget(panel)
        layout.addWidget(center_container, alignment=Qt.AlignCenter)

        self.setLayout(layout)
        self.setStyleSheet("background-color: #dcdcdc;")


class HelpWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setMinimumSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        # Header
        self.header = CustomHeader(self, active="Help")
        layout.addWidget(self.header)

        # Title
        title = QLabel("Input error code and click «Enter» button")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 14px; font-weight: bold;")
        layout.addWidget(title)

        # Main panel (flexible)
        panel = QFrame()
        panel.setStyleSheet("background-color: #bbbbbb; border: 2px solid black;")
        panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(15, 10, 15, 10)
        panel_layout.setSpacing(10)

        # Input field and button
        # Input field and buttons (Enter + Home)
        input_row = QHBoxLayout()
        self.input_field = QLineEdit()
        self.input_field.setPlaceholderText("Input Error Code")
        self.input_field.setFixedHeight(28)
        self.input_field.setStyleSheet("background-color: white; font-size: 13px; padding: 4px; QLineEdit::placeholder { font-weight: bold;};")
        self.input_field.setFocus()

        self.enter_btn = QPushButton("Enter")
        self.enter_btn.setFixedSize(80, 28)
        self.enter_btn.setStyleSheet(self.button_style(font_size="13px", bold=True))
        self.enter_btn.clicked.connect(self.perform_search)

        self.home_btn = QPushButton("Home")
        self.home_btn.setFixedSize(80, 28)
        self.home_btn.setStyleSheet(self.button_style(font_size="13px", bold=True))
        self.home_btn.clicked.connect(self.back_to_home)

        input_row.addWidget(self.input_field)
        input_row.addWidget(self.enter_btn)
        input_row.addWidget(self.home_btn)
        panel_layout.addLayout(input_row)


        # Result row
        result_row = QHBoxLayout()

        # Left side - Cause
        cause_box = QVBoxLayout()
        cause_label = QLabel("Cause")
        cause_label.setAlignment(Qt.AlignCenter)
        cause_label.setStyleSheet("font-weight: bold; font-size: 13px; margin: 1px; padding: 1px;")
        cause_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        self.cause_result = QLabel()
        self.cause_result.setStyleSheet("""
            background-color: #f0f0f0;
            border: 1px solid #bbb;
            padding: 6px;
            font-size: 13px;
        """)
        self.cause_result.setWordWrap(True)
        self.cause_result.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        cause_box.addWidget(cause_label)
        cause_box.addWidget(self.cause_result)

        # Right side - Corrective Actions
        action_box = QVBoxLayout()
        action_label = QLabel("Corrective Actions")
        action_label.setAlignment(Qt.AlignCenter)
        action_label.setStyleSheet("font-weight: bold; font-size: 13px; margin: 1px; padding: 1px;")
        action_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        self.action_result = QLabel()
        self.action_result.setStyleSheet("""
            background-color: #f0f0f0;
            border: 1px solid #bbb;
            padding: 6px;
            font-size: 13px;
        """)
        self.action_result.setWordWrap(True)
        self.action_result.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        action_box.addWidget(action_label)
        action_box.addWidget(self.action_result)

        result_row.addLayout(cause_box)
        result_row.addLayout(action_box)
        panel_layout.addLayout(result_row)

        layout.addWidget(panel)
        self.setLayout(layout)
    
    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            self.perform_search()

    def perform_search(self):
        code = self.input_field.text().strip()

        self.cause_result.clear()
        self.action_result.clear()

        if not code:
            QMessageBox.warning(self, "Input Error", "Please enter an error code.")
            return

        if not self.df:
            QMessageBox.critical(self, "Data Error", "There is no mapping table under the Result directory.")
            return

        # Find the first line by error
        target_row = next((row for row in self.df if row["Err Code"] == code), None)

        if not target_row:
            self.cause_result.setText("We didn't find the mathces with the searched error code")
            self.action_result.setText("We didn't find the mathces with the searched error code")
            return

        cause_value = target_row["Cause"] or "None"

        # Find all lines with same cause
        matched_actions = set()
        for row in self.df:
            if row["Cause"] == cause_value and row["Action"]:
                matched_actions.add(row["Action"])

        actions_text = "\n".join(sorted(matched_actions))

        self.cause_result.setText(cause_value)
        self.action_result.setText(actions_text if actions_text else "No corrective action available.")

    def back_to_home(self):
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()


class AboutWindow(BaseWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setMinimumSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        # Custom header with highlight
        self.header = CustomHeader(self, active="About")
        layout.addWidget(self.header)

        # Central container
        center_container = QWidget()
        center_layout = QVBoxLayout(center_container)
        center_layout.setAlignment(Qt.AlignCenter)

        panel = QFrame()
        panel.setStyleSheet("""
            background-color: #f0f0f0;
            border: 1px solid #999999;
            border-radius: 10px;
        """)
        panel.setFixedSize(640, 420)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(30, 25, 30, 25)
        panel_layout.setSpacing(15)

        # Title
        title = QLabel("About the Application")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #333333;")
        panel_layout.addWidget(title)

        # App Description
        description = QLabel(
            f"This system provides intelligent log file analysis and error mapping support. Current version of the mapping table is {BaseWindow.latest_version:04d}.\n"
            "It helps users identify issues by searching error codes and reviewing log contexts."
        )
        description.setAlignment(Qt.AlignCenter)
        description.setWordWrap(True)
        description.setStyleSheet("font-size: 14px; color: #222222;")
        panel_layout.addWidget(description)

        # Team Section
        team_title = QLabel("Development Team")
        team_title.setAlignment(Qt.AlignCenter)
        team_title.setStyleSheet("font-size: 16px; font-weight: bold; margin-top: 10px;")
        panel_layout.addWidget(team_title)

        # Team Members
        team_members = [
            "Rachel – Project Leader, Tester",
            "Aleksandr – Graphical Interface Designer, Tester",
            "Mubashir – Low Level Designer, Support",
            "Aleks – Team Leader, Developer"
        ]

        for member in team_members:
            member_label = QLabel(member)
            member_label.setAlignment(Qt.AlignLeft)
            member_label.setStyleSheet("font-size: 13px; color: #111111; padding-left: 10px;")
            panel_layout.addWidget(member_label)


        # Home button inside the panel
        home_btn = QPushButton("Home")
        home_btn.setFixedHeight(35)
        home_btn.setStyleSheet(self.button_style(font_size="14px", bold=True))
        home_btn.clicked.connect(self.back_to_home)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(home_btn)
        btn_row.addStretch()

        panel_layout.addLayout(btn_row)

        center_layout.addWidget(panel)
        layout.addWidget(center_container, alignment=Qt.AlignCenter)
        self.setLayout(layout)

    def back_to_home(self):
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()

class UserChoiceWindow(BaseWindow):
    def __init__(self, selected_files):
        super().__init__()
        self.selected_files = selected_files
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setFixedSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(20)

        self.header = CustomHeader(self, active="Search")
        main_layout.addWidget(self.header)

        center_container = QWidget()
        center_layout = QVBoxLayout(center_container)
        center_layout.setAlignment(Qt.AlignCenter)

        panel = QFrame()
        panel.setStyleSheet("background-color: #bbbbbb; border: 2px solid gray;")
        panel.setFixedSize(300, 250)

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(20, 20, 20, 20)
        panel_layout.setSpacing(20)

        # Buttons
        self.manual_btn = QPushButton("Manual")
        self.auto_btn = QPushButton("Auto")
        self.home_btn = QPushButton("Home")

        for btn in [self.manual_btn, self.auto_btn, self.home_btn]:
            btn.setFixedHeight(40)
            btn.setStyleSheet(self.button_style(font_size="16px", bold=True))

            panel_layout.addWidget(btn)

        center_layout.addWidget(panel)
        main_layout.addWidget(center_container, alignment=Qt.AlignCenter)

        self.setLayout(main_layout)

        # Logics
        self.manual_btn.clicked.connect(self.open_manual)
        self.auto_btn.clicked.connect(self.open_auto)
        self.home_btn.clicked.connect(self.back_to_main)

    def open_manual(self):
        self.manual_window = ManualModeWindow(self.selected_files)
        self.manual_window.show()
        self.close()

    def open_auto(self):
        self.analysis_window = AnalyzingWindow(self.selected_files, "EALM")
        self.analysis_window.show()
        self.close()


    def back_to_main(self):
        self.selected_files.clear()
        self.close()
        self.main_window = MainWindow()
        self.main_window.show()

class ManualModeWindow(BaseWindow):
    def __init__(self, selected_files):
        super().__init__()
        self.selected_files = selected_files
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setFixedSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(20)

        self.header = CustomHeader(self, active="Search")
        main_layout.addWidget(self.header)

        center_container = QWidget()
        center_layout = QVBoxLayout(center_container)
        center_layout.setAlignment(Qt.AlignCenter)

        panel = QFrame()
        panel.setFixedSize(400, 250)
        panel.setStyleSheet("background-color: #bbbbbb; border: 2px solid gray;")

        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(20, 20, 20, 20)
        panel_layout.setSpacing(20)

        # Input
        self.input_field = QLineEdit()
        self.input_field.setPlaceholderText("Input Search Text") 
        self.input_field.setStyleSheet("font-size: 16px; padding: 5px;")
        self.input_field.setMaxLength(255)
        panel_layout.addWidget(self.input_field)


        # Buttons
        button_row = QHBoxLayout()
        button_names = [("Start Search", self.start_search),
                        ("Back", self.go_back),
                        ("Home", self.go_home)]

        for name, handler in button_names:
            btn = QPushButton(name)
            btn.setFixedHeight(35)
            btn.setStyleSheet(self.button_style(font_size="16px", bold=True))
            btn.clicked.connect(handler)
            button_row.addWidget(btn)

        panel_layout.addLayout(button_row)
        center_layout.addWidget(panel)
        main_layout.addWidget(center_container, alignment=Qt.AlignCenter)

        self.setLayout(main_layout)

    def start_search(self):
        search_text = self.input_field.text().strip()

        if not search_text:
            QMessageBox.warning(self, "Input Error", "Please enter the text you want to seacrh.")
            return

        self.analysis_window = AnalyzingWindow(self.selected_files, search_text)
        self.analysis_window.show()
        self.close()


    def go_back(self):
        self.user_choice_window = UserChoiceWindow(self.selected_files)
        self.user_choice_window.show()
        self.close()

    def go_home(self):
        self.selected_files.clear()
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()

class AnalyzingWindow(BaseWindow):
    def __init__(self, selected_files, search_text):
        super().__init__()
        self.selected_files = selected_files
        self.search_text = search_text

        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setFixedSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(20)

        self.header = CustomHeader(self, active="Search")
        self.layout.addWidget(self.header)

        self.label = QLabel("Analyzing...")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("font-size: 24px; font-weight: bold;")
        self.layout.addWidget(self.label, alignment=Qt.AlignCenter)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setFixedWidth(400)
        self.layout.addWidget(self.progress, alignment=Qt.AlignCenter)

        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.setFixedHeight(35)
        self.cancel_btn.setStyleSheet("background-color: white; border: 1px solid black; font-size: 14px;")
        self.cancel_btn.clicked.connect(self.cancel_analysis)
        self.layout.addWidget(self.cancel_btn, alignment=Qt.AlignCenter)

        self.setLayout(self.layout)

        # Analyze the data
        self.result_data = []
        self.total_lines = sum(1 for f in selected_files for _ in open(f, encoding='utf-8', errors='ignore'))
        self.current_line = 0

        self.file_iter = iter(self.selected_files)
        self.current_file = None
        self.lines_iter = iter([])
        self.line_num = 0
        self.skip_ahead = 0

        self.timer = QTimer()
        self.timer.timeout.connect(self.step_analysis)
        self.timer.start(1)



    def step_analysis(self):
        """Step-by-step log file analysis with case-insensitive search based on user input."""

        try:
            # If no file is currently loaded, load the next one
            if self.current_file is None:
                file = next(self.file_iter)
                self.current_file = file
                with open(file, 'r', encoding='utf-8', errors='ignore') as f:
                    self.lines = f.readlines()
                self.lines_iter = iter(self.lines)
                self.line_num = 0

            # Read the next line from the current file
            line = next(self.lines_iter)
            self.line_num += 1
            self.current_line += 1

            # Normalize case for search
            line_lower = line.lower()
            search_input = self.search_text.strip().lower()

            search_input = self.search_text.strip().lower()
            line_lower = line.lower()

            # Divide on termins 
            if ',' in search_input or ';' in search_input:
                search_terms = {term for term in re.split(r'[;,]', search_input) if term}
                matched = any(term in line_lower for term in search_terms)
            else:
                matched = search_input in line_lower


            if matched:
                self.result_data.append({
                    "file": os.path.basename(self.current_file),
                    "line": self.line_num,
                    "text": line.strip(),
                    "path": self.current_file
                })

            # Update progress bar
            progress = int((self.current_line / self.total_lines) * 100)
            self.progress.setValue(progress)

        except StopIteration:
            self.current_file = None
            try:
                next_file = next(self.file_iter)
                self.current_file = next_file
                with open(next_file, 'r', encoding='utf-8', errors='ignore') as f:
                    self.lines = f.readlines()
                self.lines_iter = iter(self.lines)
                self.line_num = 0
                self.step_analysis()
            except StopIteration:
                self.timer.stop()
                self.open_result()


    def open_result(self):
        if not self.result_data:
            # No results found — open the "Nothing Found" window
            self.nothing_window = NothingFoundWindow(self.selected_files)
            self.nothing_window.show()
        else:
            # Results found — open the results window
            self.result_window = FoundResultWindow(self.result_data, self.selected_files)
            self.result_window.show()
        self.close()


    def cancel_analysis(self):
        self.timer.stop()
        self.user_choice_window = UserChoiceWindow(self.selected_files)
        self.user_choice_window.show()
        self.close()


class NothingFoundWindow(BaseWindow):
    def __init__(self, selected_files):
        super().__init__()
        self.selected_files = selected_files
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setFixedSize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        # Main layout for the whole window
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        # Custom header at the top
        self.header = CustomHeader(self, active="Search")
        layout.addWidget(self.header)

        # Center container with fixed size
        center_widget = QWidget()
        center_layout = QVBoxLayout(center_widget)
        center_layout.setAlignment(Qt.AlignCenter)

        panel = QFrame()
        panel.setFixedSize(300, 180)
        panel.setStyleSheet("background-color: #bbbbbb; border: 2px solid gray;")
        panel_layout = QVBoxLayout(panel)
        panel_layout.setContentsMargins(20, 20, 20, 20)
        panel_layout.setSpacing(20)

        # Label
        label = QLabel("We didn't find the mathces with the searched text")
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("font-size: 16px; font-weight: bold;")
        panel_layout.addWidget(label)

        # Buttons row
        btn_layout = QHBoxLayout()
        self.back_btn = QPushButton("Back")
        self.home_btn = QPushButton("Home")

        for btn in (self.back_btn, self.home_btn):
            btn.setFixedHeight(35)
            btn.setStyleSheet(self.button_style(font_size="14px", bold=True))
            btn_layout.addWidget(btn)

        panel_layout.addLayout(btn_layout)
        center_layout.addWidget(panel)

        layout.addWidget(center_widget, alignment=Qt.AlignCenter)
        self.setLayout(layout)

        # Button logic
        self.back_btn.clicked.connect(self.back_to_selection)
        self.home_btn.clicked.connect(self.back_to_home)

    def back_to_selection(self):
        self.user_choice_window = UserChoiceWindow(self.selected_files)
        self.user_choice_window.show()
        self.close()

    def back_to_home(self):
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()


class FoundResultWindow(BaseWindow):
    def __init__(self, results, selected_files):
        super().__init__()
        self.results = results
        self.selected_files = selected_files

        self.setWindowFlags(Qt.FramelessWindowHint)
        self.resize(800, 600)
        self.setStyleSheet("background-color: #dcdcdc;")

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        self.header = CustomHeader(self, active="Search")
        layout.addWidget(self.header)

        info = QLabel("Select «Text in a Row» to see log file")
        info.setAlignment(Qt.AlignCenter)
        info.setStyleSheet("font-weight: bold; font-size: 14px;")
        layout.addWidget(info)

        # Button for extract the table's result 
        self.expand_table_btn = QPushButton("🔳")
        self.expand_table_btn.setFixedSize(24, 24)
        self.expand_table_btn.clicked.connect(self.toggle_table_size)

        # Button for extract the log file's lines
        self.expand_log_btn = QPushButton("🔳")
        self.expand_log_btn.setFixedSize(24, 24)
        self.expand_log_btn.clicked.connect(self.toggle_log_size)

        # Table of results
        self.result_area = QTableWidget()
        self.result_area.setMaximumHeight(200)
        self.result_area.setColumnCount(3)
        self.result_area.setHorizontalHeaderLabels(["File Name", "Location", "Text in a Row"])
        self.result_area.setRowCount(len(self.results))
        self.result_area.setStyleSheet("background-color: white;")
        self.result_area.setEditTriggers(QTableWidget.NoEditTriggers)
        self.result_area.setSelectionBehavior(QTableWidget.SelectRows)

        header = self.result_area.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.Stretch)

        for row, res in enumerate(self.results):
            self.result_area.setItem(row, 0, QTableWidgetItem(res["file"]))
            self.result_area.setItem(row, 1, QTableWidgetItem(f"Line {res['line']}"))
            self.result_area.setItem(row, 2, QTableWidgetItem(res["text"]))

        self.result_area.cellClicked.connect(self.show_log_context)

        table_container = QHBoxLayout()
        table_container.addWidget(self.result_area)
        table_container.addWidget(self.expand_table_btn, alignment=Qt.AlignTop)
        layout.addLayout(table_container)


        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setStyleSheet("background-color: #eeeeee; font-family: monospace; font-size: 12px;")
        self.log_output.setMinimumHeight(250)
        
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setWidget(self.log_output)
        self.scroll.setMaximumHeight(250)

        log_container = QHBoxLayout()
        log_container.addWidget(self.scroll)
        log_container.addWidget(self.expand_log_btn, alignment=Qt.AlignTop)
        layout.addLayout(log_container)

        if self.result_area.maximumHeight() < 500:
            self.result_area.setMaximumHeight(800)
        else:
            self.result_area.setMaximumHeight(200)

        if self.scroll.maximumHeight() < 500:
            self.scroll.setMaximumHeight(800)
        else:
            self.scroll.setMaximumHeight(250)


        # Buttons
        btn_row = QHBoxLayout()
        self.back_btn = QPushButton("Back")
        self.back_btn.clicked.connect(self.back)

        self.home_btn = QPushButton("Home")
        self.home_btn.clicked.connect(self.go_home)

        for btn in [self.back_btn, self.home_btn]:
            btn.setFixedHeight(35)
            btn.setStyleSheet(self.button_style(font_size="14px", bold=True))
            btn_row.addWidget(btn)

        layout.addLayout(btn_row)
        self.setLayout(layout)

    def show_log_context(self, row):
        line_str = self.result_area.item(row, 1).text()
        try:
            line_num = int(line_str.replace("Line", "").strip())
        except ValueError:
            self.log_output.setText("Invalid line number format.")
            return

        file_path = self.results[row].get("path")
        if not file_path or not os.path.exists(file_path):
            self.log_output.setText("File not found.")
            return

        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                lines = f.readlines()

            total_lines = len(lines)
            center_index = line_num

            # 250 above and below
            start = max(0, center_index - 250)
            end = min(total_lines, center_index + 251)
            context = lines[start:end]

            # Format and display context
            display_lines = []
            for i, line in enumerate(context, start=start + 1):
                prefix = f">>> Line {i}: " if i == center_index else f"Line {i}: "
                display_lines.append(prefix + line.rstrip())

            full_text = "\n".join(display_lines)
            self.log_output.setPlainText(full_text)

            # Highlight line
            highlight_line = f">>> Line {center_index}:"
            text = self.log_output.toPlainText()
            pos = text.find(highlight_line)
            if pos != -1:
                cursor = self.log_output.textCursor()
                cursor.setPosition(pos)
                cursor.movePosition(QTextCursor.EndOfLine, QTextCursor.KeepAnchor)

                fmt = QTextCharFormat()
                fmt.setBackground(QColor("yellow"))
                fmt.setFontWeight(QFont.Bold)
                cursor.mergeCharFormat(fmt)

                # Center scroll to highlighted line
                block_number = text[:pos].count("\n")
                block_height = self.log_output.fontMetrics().height()
                visible_blocks = int(self.log_output.viewport().height() / block_height)
                center_scroll_value = max(0, block_number - visible_blocks // 2)
                self.log_output.verticalScrollBar().setValue(center_scroll_value)

        except Exception as e:
            self.log_output.setPlainText(f"Error reading file: {e}")


    def toggle_table_size(self):
        self.popup = ResultTablePopup(self.result_area)
        self.popup.showMaximized()

    def toggle_log_size(self):
        searched_text = ""
        current_row = self.result_area.currentRow()
        if current_row >= 0:
            searched_text = self.results[current_row].get("text", "").strip()

        self.log_popup = LogFilePopup(self.log_output.toPlainText(), searched_text)
        self.log_popup.showMaximized()

    def back(self):
        self.user_choice_window = UserChoiceWindow(self.selected_files)
        self.user_choice_window.show()
        self.close()

    def go_home(self):
        self.main_window = MainWindow()
        self.main_window.show()
        self.close()

class ResultTablePopup(QWidget):
    def __init__(self, table_widget):
        super().__init__()
        self.setWindowTitle("Full Result Table")
        self.resize(800, 600)
        layout = QVBoxLayout(self)

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        table_copy = QTableWidget()
        table_copy.setColumnCount(table_widget.columnCount())
        table_copy.setRowCount(table_widget.rowCount())
        table_copy.setHorizontalHeaderLabels([table_widget.horizontalHeaderItem(i).text() for i in range(table_widget.columnCount())])
        table_copy.setEditTriggers(QTableWidget.NoEditTriggers)
        table_copy.setSelectionBehavior(QTableWidget.SelectRows)

        for row in range(table_widget.rowCount()):
            for col in range(table_widget.columnCount()):
                item = table_widget.item(row, col)
                table_copy.setItem(row, col, QTableWidgetItem(item.text() if item else ""))

        layout.addWidget(table_copy)

class LogFilePopup(QWidget):
    def __init__(self, log_text, highlight_text=""):
        super().__init__()
        self.setWindowTitle("Full Log View")
        self.resize(800, 600)
        layout = QVBoxLayout(self)

        if BaseWindow.window_is_maximized:
            self.showMaximized()
        else:
            self.showNormal()

        self.text_area = QTextEdit()
        self.text_area.setReadOnly(True)
        self.text_area.setStyleSheet("background-color: #eeeeee; font-family: monospace; font-size: 12px;")
        self.text_area.setPlainText(log_text)
        layout.addWidget(self.text_area)

        if highlight_text:
            self.highlight_text_occurrences(highlight_text)

    def highlight_text_occurrences(self, keyword):
        cursor = self.text_area.textCursor()
        fmt = QTextCharFormat()
        fmt.setBackground(QColor("yellow"))
        fmt.setFontWeight(QFont.Bold)

        content = self.text_area.toPlainText()
        pattern = re.escape(keyword)
        regex = re.compile(pattern, re.IGNORECASE)

        for match in regex.finditer(content):
            start, end = match.span()
            cursor.setPosition(start)
            cursor.setPosition(end, QTextCursor.KeepAnchor)
            cursor.setCharFormat(fmt)

        # Go to first matching
        first = regex.search(content)
        if first:
            cursor.setPosition(first.start())
            self.text_area.setTextCursor(cursor)
            self.text_area.ensureCursorVisible()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
